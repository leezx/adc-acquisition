"""Normalize EMA's bulk medicines XLSX and per-medicine EPAR HTML page.

The medicines XLSX (https://www.ema.europa.eu/en/medicines/download-medicine-data)
is a static export with 8 metadata rows before the real header row, then
one row per authorised/refused/withdrawn medicine — verified live on
2026-08-12. The EPAR page per medicine is plain static HTML with a
repeating "language card" per document (one card per translation);
English document links live inside a card whose language label is
"English (EN)", each with its own "Last updated" timestamp. Defensive
throughout: a missing column/date/card must never crash the batch.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass

import openpyxl

HEADER_ROW_INDEX = 8  # 0-indexed; first 8 rows are report metadata, not data.

# Column indices in the medicines XLSX, verified live on 2026-08-12.
COL_NAME = 1
COL_PRODUCT_NUMBER = 2
COL_STATUS = 3
COL_ACTIVE_SUBSTANCE = 7
COL_THERAPEUTIC_AREA = 8
COL_MAH = 25
COL_DECISION_DATE = 26
COL_AUTHORISATION_DATE = 31
COL_WITHDRAWAL_DATE = 33
COL_LAST_UPDATED = 37
COL_URL = 38

# Real columns end at 38; openpyxl's read_only iterator otherwise pads
# rows out with hundreds of trailing None cells from the sheet's declared
# dimensions.
REAL_COLUMN_COUNT = 39


@dataclass
class ParsedMedicine:
    product_number: str
    name: str | None
    status: str | None
    active_substance: str | None
    therapeutic_area: str | None
    marketing_authorisation_holder: str | None
    decision_date: str | None  # normalized YYYY-MM-DD
    authorisation_date: str | None
    withdrawal_date: str | None
    last_updated_date: str | None
    epar_url: str | None
    raw_row: dict  # every real column, header-name -> raw cell value, for full-fidelity raw storage


@dataclass
class EparDocument:
    filename: str
    doc_type: str | None
    url: str
    last_updated: str | None  # normalized YYYY-MM-DD


_UK_DATE_RE = re.compile(r"^(\d{2})/(\d{2})/(\d{4})$")


def normalize_ema_date(raw) -> str | None:
    """EMA's XLSX dates are DD/MM/YYYY strings; the universal manifest
    contract wants YYYY-MM-DD."""
    if not raw or not isinstance(raw, str):
        return None
    m = _UK_DATE_RE.match(raw.strip())
    if not m:
        return None
    day, month, year = m.groups()
    return f"{year}-{month}-{day}"


def parse_medicines_xlsx(xlsx_bytes: bytes) -> list[ParsedMedicine]:
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header: list[str] = []
    medicines = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < HEADER_ROW_INDEX:
            continue
        if i == HEADER_ROW_INDEX:
            header = [str(h) if h is not None else f"col_{j}" for j, h in enumerate(row[:REAL_COLUMN_COUNT])]
            continue
        product_number = row[COL_PRODUCT_NUMBER] if len(row) > COL_PRODUCT_NUMBER else None
        if not product_number:
            continue
        medicines.append(
            ParsedMedicine(
                product_number=product_number,
                name=row[COL_NAME] if len(row) > COL_NAME else None,
                status=row[COL_STATUS] if len(row) > COL_STATUS else None,
                active_substance=row[COL_ACTIVE_SUBSTANCE] if len(row) > COL_ACTIVE_SUBSTANCE else None,
                therapeutic_area=row[COL_THERAPEUTIC_AREA] if len(row) > COL_THERAPEUTIC_AREA else None,
                marketing_authorisation_holder=row[COL_MAH] if len(row) > COL_MAH else None,
                decision_date=normalize_ema_date(row[COL_DECISION_DATE]) if len(row) > COL_DECISION_DATE else None,
                authorisation_date=normalize_ema_date(row[COL_AUTHORISATION_DATE]) if len(row) > COL_AUTHORISATION_DATE else None,
                withdrawal_date=normalize_ema_date(row[COL_WITHDRAWAL_DATE]) if len(row) > COL_WITHDRAWAL_DATE else None,
                last_updated_date=normalize_ema_date(row[COL_LAST_UPDATED]) if len(row) > COL_LAST_UPDATED else None,
                epar_url=row[COL_URL] if len(row) > COL_URL else None,
                raw_row=dict(zip(header, row[:REAL_COLUMN_COUNT])),
            )
        )
    wb.close()
    return medicines


def is_adc_candidate(medicine: ParsedMedicine, substance_patterns: list[str]) -> bool:
    """Systematic INN-suffix matching (vedotin, emtansine, deruxtecan, ...
    — standardized WHO INN stems for ADC linker/payload chemistry), not a
    manually maintained list of specific approved drug names — this
    catches future ADCs using the same naming convention without needing
    per-drug updates, same spirit as Job 06 (FDA)'s full-text
    "antibody-drug conjugate" label search."""
    haystack = f"{medicine.name or ''} {medicine.active_substance or ''}".lower()
    return any(pattern.lower() in haystack for pattern in substance_patterns)


_DOC_LANGUAGE_CARD_RE = re.compile(
    r'<p class="language-meta[^"]*"[^>]*>English \(EN\).*?'
    r'<time datetime="([^"]+)"[^>]*>.*?'
    r'<a[^>]*href="([^"]+_en\.pdf)"',
    re.S,
)
_DOC_TYPE_RE = re.compile(r"/en/documents/([^/]+)/")


def parse_epar_documents(html: str, base_url: str = "https://www.ema.europa.eu") -> list[EparDocument]:
    """English-language PDF documents linked from a medicine's EPAR page
    (Prompt.md's "EPAR, product information, assessment reports,
    authorisation history, safety updates" — the actual PDFs, not the
    structured metadata already covered by the medicine's own manifest
    row)."""
    documents = []
    seen_urls = set()
    for last_updated_raw, path in _DOC_LANGUAGE_CARD_RE.findall(html):
        url = path if path.startswith("http") else f"{base_url}{path}"
        if url in seen_urls:
            continue
        seen_urls.add(url)
        type_match = _DOC_TYPE_RE.search(path)
        documents.append(
            EparDocument(
                filename=path.rsplit("/", 1)[-1],
                doc_type=type_match.group(1) if type_match else None,
                url=url,
                last_updated=last_updated_raw[:10] if last_updated_raw else None,  # already YYYY-MM-DD in the ISO datetime
            )
        )
    return documents


def within_date_range(date_str: str | None, since: str | None, until: str | None) -> bool:
    """Lexicographic YYYY-MM-DD comparison — same convention as
    jobs/sec/parser.py and jobs/fda/parser.py."""
    if not since and not until:
        return True
    if not date_str:
        return False
    if since and date_str < since:
        return False
    if until and date_str > until:
        return False
    return True
